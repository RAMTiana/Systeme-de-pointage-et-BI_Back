"""
Rendu Excel des rapports (connecteur d'export openpyxl).
Extrait de `rapport_service` pour respecter la limite de 500 lignes par fichier.
"""
import os
import re
from datetime import date as date_
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImageOpenpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas as canvas_mod
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from app.core import jours_feries
from app.core.config import settings
from app.models.agent import Agent
from app.models.anomalie import Anomalie
from app.models.conge import Conge
from app.models.enums import FormatRapport, JourSemaine, StatutAgent, StatutConge, StatutPointage, TypeAnomalie, \
    TypePeriode, TypePointage
from app.models.horaire_reference import HoraireReference
from app.models.pointage import Pointage
from app.models.rapport import Rapport
from app.models.service import Service
from app.services import horaire_service, journal_audit_service

from app.services.rapport_commun import (
    _AMBRE,
    _AMBRE_HEX,
    _AMBRE_LIGHT_HEX,
    _BANNIERE_H,
    _BLANC,
    _BLEU,
    _BLEU_HEX,
    _CHEMIN_LOGO,
    _CORAIL,
    _CORAIL_HEX,
    _CORAIL_LIGHT_HEX,
    _GRIS_BORD,
    _GRIS_BORD_HEX,
    _GRIS_ZEBRE,
    _GRIS_ZEBRE_HEX,
    _INFO_LIGHT_HEX,
    _JOURS_PAR_INDEX,
    _LIBELLE_PERIODE,
    _MARGE,
    _NOM_ORG,
    _PAGE_H,
    _PAGE_W,
    _PIED_H,
    _SOUS_NOM_ORG,
    _TEAL,
    _TEAL_HEX,
    _TEAL_LIGHT_HEX,
    _TEXTE,
    _TEXTE_HEX,
    _TEXTE_MUT,
    _TEXTE_MUT_HEX,
    _hex_argb,
    _logo_disponible,
)


def _fill(couleur_hex: str) -> PatternFill:
    return PatternFill(start_color=couleur_hex, end_color=couleur_hex, fill_type="solid")


_BORDURE_FINE = Border(*(Side(style="thin", color=_GRIS_BORD_HEX) for _ in range(4)))
_POLICE_ENTETE = Font(bold=True, color="FFFFFF", size=9.5)
_FOND_ENTETE = _fill(_BLEU_HEX)
_FOND_ZEBRE = _fill(_GRIS_ZEBRE_HEX)
_ALIGN_CENTRE = Alignment(horizontal="center", vertical="center")
_ALIGN_GAUCHE = Alignment(horizontal="left", vertical="center")
_FORMAT_POURCENT = "[$-40C]0.0%"
_FORMAT_HEURES = "[$-40C]#,##0.0 \"h\""
_FORMAT_ENTIER = "[$-40C]#,##0"

# Accent + fond pastel par carte KPI, dans l'ordre d'affichage (Agents,
# Taux — calculé dynamiquement, Retards, Absences, Départs anticipés,
# Heures). Reprend les couleurs de la palette et les teintes --*-light du
# frontend pour que les cartes KPI ressemblent à leur équivalent à l'écran.
_ACCENTS_KPI = [_BLEU_HEX, None, _AMBRE_HEX, _CORAIL_HEX, _AMBRE_HEX, _BLEU_HEX]
_FONDS_KPI = [_INFO_LIGHT_HEX, None, _AMBRE_LIGHT_HEX, _CORAIL_LIGHT_HEX, _AMBRE_LIGHT_HEX, _INFO_LIGHT_HEX]


def _couleur_taux_hex(taux: Optional[float]) -> str:
    if taux is None:
        return _TEXTE_MUT_HEX
    if taux >= 0.90:
        return _TEAL_HEX
    if taux >= 0.75:
        return _AMBRE_HEX
    return _CORAIL_HEX


def _couleur_taux_claire_hex(taux: Optional[float]) -> str:
    if taux is None:
        return _GRIS_ZEBRE_HEX
    if taux >= 0.90:
        return _TEAL_LIGHT_HEX
    if taux >= 0.75:
        return _AMBRE_LIGHT_HEX
    return _CORAIL_LIGHT_HEX


def _configurer_impression(ws, nb_colonnes: int, titre_pied: str) -> None:
    """Mise en page impression standard (paysage, ajustée à la largeur, en-tête répété, pied de page)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 1.2
    ws.page_margins.top = 1.4
    ws.page_margins.bottom = 1.2
    ws.print_title_rows = "1:1"
    ws.oddHeader.center.text = titre_pied
    ws.oddHeader.center.size = 9
    ws.oddFooter.left.text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — document à usage interne"
    ws.oddFooter.right.text = "Page &P / &N"
    ws.oddFooter.left.size = ws.oddFooter.right.size = 8


def _ecrire_feuille_detail(
    wb: Workbook, nom_feuille: str, entetes: List[str], lignes: List[list],
    idx_col_taux: Optional[int], titre_pied: str,
):
    """Construit une feuille de détail standardisée : en-tête coloré, bordures,
    zébrage, colonne 'Taux de présence' au format pourcentage et code couleur
    identique au PDF, gel de l'en-tête, filtre automatique, mise en page
    impression prête à l'emploi."""
    ws = wb.create_sheet(nom_feuille)
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = _POLICE_ENTETE
        cell.fill = _FOND_ENTETE
        cell.alignment = _ALIGN_CENTRE
        cell.border = _BORDURE_FINE
    ws.row_dimensions[1].height = 20

    for i, ligne in enumerate(lignes, start=2):
        for col, valeur in enumerate(ligne, start=1):
            cell = ws.cell(row=i, column=col, value=valeur)
            cell.border = _BORDURE_FINE
            cell.alignment = _ALIGN_CENTRE if col > 1 else _ALIGN_GAUCHE
            if i % 2 == 0:
                cell.fill = _FOND_ZEBRE
            if isinstance(valeur, int) and not isinstance(valeur, bool) and col != idx_col_taux:
                cell.number_format = _FORMAT_ENTIER
            if idx_col_taux is not None and col == idx_col_taux:
                if isinstance(valeur, (int, float)):
                    cell.number_format = _FORMAT_POURCENT
                cell.font = Font(bold=True, color=_couleur_taux_hex(valeur if isinstance(valeur, (int, float)) else None))

    # Largeur de colonne ajustée au contenu réel (borne pour éviter les
    # colonnes démesurées sur un libellé exceptionnellement long).
    for col in range(1, len(entetes) + 1):
        lettre = get_column_letter(col)
        largeur_max = max(
            [len(str(entetes[col - 1]))] + [len(str(ligne[col - 1])) for ligne in lignes if ligne[col - 1] is not None]
        )
        ws.column_dimensions[lettre].width = min(max(largeur_max + 3, 12), 32)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(entetes))}{len(lignes) + 1}"
    ws.sheet_view.showGridLines = False
    _configurer_impression(ws, len(entetes), titre_pied)
    return ws


def _rendre_excel(chemin_absolu: str, indicateurs: dict) -> None:
    libelle_periode = _LIBELLE_PERIODE[indicateurs["type_periode"]]
    periode_txt = (
        f"Du {indicateurs['periode_debut'].strftime('%d/%m/%Y')} "
        f"au {indicateurs['periode_fin'].strftime('%d/%m/%Y')}"
    )
    titre_pied = f"Rapport {libelle_periode} de présence — {indicateurs['nom_service']}"

    wb = Workbook()

    # Propriétés du document (norme de gestion documentaire : titre, auteur,
    # sujet, société — visibles dans les propriétés du fichier, utiles pour
    # l'archivage et la recherche).
    wb.properties.title = titre_pied
    wb.properties.subject = f"Rapport de présence {libelle_periode}, période du " \
        f"{indicateurs['periode_debut'].isoformat()} au {indicateurs['periode_fin'].isoformat()}"
    wb.properties.creator = "Système de pointage et BI"
    wb.properties.description = _NOM_ORG
    wb.properties.category = "Rapport de présence"
    wb.properties.keywords = "présence, pointage, SRB, rapport"

    # ---- Feuille "Synthèse" ----
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _BLEU_HEX

    # Bandeau institutionnel (fond bleu marque + liséré teal), même
    # traitement que l'en-tête du PDF : logo à gauche, titre/sous-titre au
    # centre, organisme à droite. Le bandeau s'étend sur une colonne de plus
    # (G) que les cartes KPI pour laisser assez de place au nom de
    # l'organisme sans qu'il ne déborde du fond coloré.
    for ligne in (1, 2):
        for col in range(1, 8):
            ws.cell(row=ligne, column=col).fill = _fill(_BLEU_HEX)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    ws.merge_cells("B1:E1")
    ws["B1"] = titre_pied
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].alignment = _ALIGN_GAUCHE

    ws.merge_cells("B2:E2")
    ws["B2"] = periode_txt
    ws["B2"].font = Font(size=9, color="D6E4EC")
    ws["B2"].alignment = _ALIGN_GAUCHE

    ws.merge_cells("F1:G1")
    ws["F1"] = _NOM_ORG
    ws["F1"].font = Font(bold=True, size=10, color="FFFFFF")
    ws["F1"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("F2:G2")
    ws["F2"] = _SOUS_NOM_ORG
    ws["F2"].font = Font(size=7.5, color="D6E4EC")
    ws["F2"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Liséré d'accent teal, identique à la bande fine sous le bandeau PDF.
    ws.row_dimensions[3].height = 4
    for col in range(1, 8):
        ws.cell(row=3, column=col).fill = _fill(_TEAL_HEX)

    if _logo_disponible():
        logo = ImageOpenpyxl(_CHEMIN_LOGO)
        logo.width = 34
        logo.height = 34
        ws.add_image(logo, "A1")

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — document à usage interne"
    ws["A4"].font = Font(size=8.5, italic=True, color=_TEXTE_MUT_HEX)
    ws.row_dimensions[4].height = 18

    g = indicateurs["globaux"]
    entetes_kpi = ["Agents concernés", "Taux de présence", "Retards", "Absences", "Départs anticipés", "Heures travaillées"]
    valeurs_kpi = [
        g["nombre_agents"],
        g["taux_presence"] if g["taux_presence"] is not None else None,
        g["nombre_retards"], g["nombre_absences"], g["nombre_departs_anticipes"], g["heures_travaillees"],
    ]

    ligne_entete_kpi, ligne_valeur_kpi = 6, 7
    for col, entete in enumerate(entetes_kpi, start=1):
        idx = col - 1
        accent_hex = _ACCENTS_KPI[idx] if _ACCENTS_KPI[idx] else _couleur_taux_hex(valeurs_kpi[1])
        fond_hex = _FONDS_KPI[idx] if _FONDS_KPI[idx] else _couleur_taux_claire_hex(valeurs_kpi[1])

        # Carte KPI = 2 lignes fusionnées visuellement par un fond pastel
        # commun, une bordure fine grise sur le pourtour et un filet coloré
        # épais en haut (le même rôle que le LINEABOVE des cartes du PDF).
        cell_entete = ws.cell(row=ligne_entete_kpi, column=col, value=entete)
        cell_entete.font = Font(bold=True, size=8, color=_TEXTE_MUT_HEX)
        cell_entete.fill = _fill(fond_hex)
        cell_entete.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        cell_entete.border = Border(
            top=Side(style="thick", color=accent_hex),
            left=Side(style="thin", color=_GRIS_BORD_HEX),
            right=Side(style="thin", color=_GRIS_BORD_HEX),
        )
        ws.column_dimensions[get_column_letter(col)].width = 20

        cell_valeur = ws.cell(row=ligne_valeur_kpi, column=col, value=valeurs_kpi[idx])
        cell_valeur.alignment = _ALIGN_CENTRE
        cell_valeur.fill = _fill(fond_hex)
        cell_valeur.font = Font(bold=True, size=14, color=_TEXTE_HEX)
        cell_valeur.border = Border(
            left=Side(style="thin", color=_GRIS_BORD_HEX),
            right=Side(style="thin", color=_GRIS_BORD_HEX),
            bottom=Side(style="thin", color=_GRIS_BORD_HEX),
        )
        if col in (1, 3, 4, 5):
            cell_valeur.number_format = _FORMAT_ENTIER
        if col == 6:
            cell_valeur.number_format = _FORMAT_HEURES

    if valeurs_kpi[1] is not None:
        cell_taux = ws.cell(row=ligne_valeur_kpi, column=2)
        cell_taux.number_format = _FORMAT_POURCENT
        cell_taux.font = Font(bold=True, size=14, color=_couleur_taux_hex(valeurs_kpi[1]))
    ws.row_dimensions[ligne_entete_kpi].height = 26
    ws.row_dimensions[ligne_valeur_kpi].height = 28

    _configurer_impression(ws, len(entetes_kpi), titre_pied)

    # ---- Feuille "Détail par service" ----
    if indicateurs["detail_services"]:
        entetes = ["Service", "Agents", "Taux de présence", "Retards", "Absences", "Départs anticipés", "Heures travaillées"]
        lignes = [
            [
                s["nom_service"], s["nombre_agents"], s["taux_presence"],
                s["nombre_retards"], s["nombre_absences"], s["nombre_departs_anticipes"], s["heures_travaillees"],
            ]
            for s in indicateurs["detail_services"]
        ]
        ws_service = _ecrire_feuille_detail(wb, "Détail par service", entetes, lignes, idx_col_taux=3, titre_pied=titre_pied)
        for i in range(2, len(lignes) + 2):
            ws_service.cell(row=i, column=7).number_format = _FORMAT_HEURES
        ws_service.sheet_properties.tabColor = _TEAL_HEX

    # ---- Feuille "Détail par agent" ----
    if indicateurs["detail_agents"]:
        entetes = ["Matricule", "Nom", "Prénom", "Taux de présence", "Retards", "Absences", "Départs anticipés", "Heures travaillées"]
        lignes = [
            [
                a["matricule"], a["nom"], a["prenom"], a["taux_presence"],
                a["nombre_retards"], a["nombre_absences"], a["nombre_departs_anticipes"], a["heures_travaillees"],
            ]
            for a in indicateurs["detail_agents"]
        ]
        ws_agent = _ecrire_feuille_detail(wb, "Détail par agent", entetes, lignes, idx_col_taux=4, titre_pied=titre_pied)
        for i in range(2, len(lignes) + 2):
            ws_agent.cell(row=i, column=8).number_format = _FORMAT_HEURES
        ws_agent.sheet_properties.tabColor = _AMBRE_HEX

    wb.save(chemin_absolu)


# --------------------------------------------------------------------
# Étape 7 : consignation en base + orchestration complète
# --------------------------------------------------------------------

